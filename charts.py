"""
画图脚本：小红书数据分析项目的 4 张核心图表
运行方式：在终端输入 python charts.py
输出：charts/ 目录下的 4 张 PNG 图片
"""

import openpyxl
import matplotlib.pyplot as plt
import matplotlib
import os
import sys

# ------------------------------
# 修复 Windows 终端 GBK 编码问题
# ------------------------------
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

# ------------------------------
# 小白必读：这 3 行是设置中文字体，否则中文字会变成方块
# ------------------------------
matplotlib.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'PingFang SC']
matplotlib.rcParams['axes.unicode_minus'] = False  # 让负号正常显示

# ------------------------------
# 第 1 步：读取数据（和之前的分析脚本一样）
# ------------------------------
filepath = r"D:\桌面\个人项目\xhs-growth-analytics\data\notes_raw.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)
ws = wb.active

notes = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    title = str(row[0]).strip() if row[0] else ''
    genre = str(row[2]).strip() if row[2] else ''
    if not title or not genre:
        continue

    exposure = float(row[3]) if row[3] else 0
    likes = float(row[6]) if row[6] else 0
    comments = float(row[7]) if row[7] else 0
    saves = float(row[8]) if row[8] else 0

    # 从发布时间提取小时
    time_str = str(row[1]) if row[1] else ''
    hour = None
    try:
        import re
        m = re.match(r'(\d{4})年(\d{2})月(\d{2})日(\d{2})时', time_str)
        if m:
            hour = int(m.group(4))
    except:
        pass

    ir = ((likes + saves + comments) / exposure * 100) if exposure > 0 else 0
    notes.append({
        'title': title, 'genre': genre,
        'exposure': exposure, 'likes': likes, 'comments': comments, 'saves': saves,
        'ir': ir, 'length': len(title), 'hour': hour,
    })
wb.close()

print(f"读取了 {len(notes)} 篇笔记的数据")

# ------------------------------
# 第 2 步：创建 charts 文件夹（如果不存在）
# ------------------------------
output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'charts')
os.makedirs(output_dir, exist_ok=True)

# 配色方案：用一组协调的颜色，看起来专业
COLOR_IMG = '#4C72B0'   # 图文：深蓝
COLOR_VID = '#DD8452'   # 视频：橙色
COLOR_HIGHLIGHT = '#55A868'  # 高亮：绿色
COLOR_SECONDARY = '#C44E52'  # 红色
PALETTE_4 = ['#4C72B0', '#55A868', '#DD8452', '#C44E52']

# ============================================================
# 图 1：视频 vs 图文 平均互动率对比
# 这是整个项目最强的发现——简单直接，一看就懂
# ============================================================
fig, ax = plt.subplots(figsize=(6, 4.5))

img_notes = [n for n in notes if n['genre'] == '图文']
vid_notes = [n for n in notes if n['genre'] == '视频']
img_ir = sum(n['ir'] for n in img_notes) / len(img_notes)
vid_ir = sum(n['ir'] for n in vid_notes) / len(vid_notes)

bars = ax.bar(
    ['图文', '视频'],
    [img_ir, vid_ir],
    color=[COLOR_IMG, COLOR_VID],
    width=0.5,
    edgecolor='white',
    linewidth=0.8,
)

# 在柱子顶部标注数值
for bar, val in zip(bars, [img_ir, vid_ir]):
    ax.text(
        bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.05,
        f'{val:.2f}%', ha='center', va='bottom', fontsize=14, fontweight='bold',
    )

# 柱子上方标注篇数
ax.text(bars[0].get_x() + bars[0].get_width() / 2, img_ir - 0.3,
        f'{len(img_notes)}篇', ha='center', va='top', fontsize=10, color='white', fontweight='bold')
ax.text(bars[1].get_x() + bars[1].get_width() / 2, vid_ir - 0.3,
        f'{len(vid_notes)}篇', ha='center', va='top', fontsize=10, color='white', fontweight='bold')

# 标注差异倍数
ax.annotate(
    f'视频是图文的 {vid_ir/img_ir:.1f} 倍',
    xy=(1, vid_ir), xytext=(1, vid_ir + 0.8),
    ha='center', fontsize=11, color=COLOR_VID,
    bbox=dict(boxstyle='round,pad=0.3', facecolor='#FFF5EE', edgecolor=COLOR_VID, alpha=0.8),
)

ax.set_ylabel('平均互动率 (%)', fontsize=12)
ax.set_title('体裁 × 互动率对比', fontsize=14, fontweight='bold', pad=15)
ax.set_ylim(0, vid_ir + 1.5)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.tick_params(axis='both', labelsize=11)

plt.tight_layout()
path = os.path.join(output_dir, '01_genre_comparison.png')
fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
plt.close(fig)
print(f'✓ 图1 已保存: {path}')

# ============================================================
# 图 2：标题字数分桶 × 互动率（倒 U 型曲线）
# 这是可操作性最强的发现——告诉你标题该写多长
# ============================================================
fig, ax = plt.subplots(figsize=(7, 4.5))

buckets = {
    '≤6 字': [n for n in notes if n['length'] <= 6],
    '7-9 字': [n for n in notes if 7 <= n['length'] <= 9],
    '10-13 字': [n for n in notes if 10 <= n['length'] <= 13],
    '≥14 字': [n for n in notes if n['length'] >= 14],
}

labels = list(buckets.keys())
ir_values = [sum(n['ir'] for n in g) / len(g) if g else 0 for g in buckets.values()]
counts = [len(g) for g in buckets.values()]

# 找到最优区间，高亮它
best_idx = ir_values.index(max(ir_values))
bar_colors = [COLOR_HIGHLIGHT if i == best_idx else '#CCCCCC' for i in range(len(labels))]

bars = ax.bar(labels, ir_values, color=bar_colors, width=0.55, edgecolor='white', linewidth=0.8)

for bar, val, cnt in zip(bars, ir_values, counts):
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.05,
            f'{val:.2f}%', ha='center', va='bottom', fontsize=13, fontweight='bold')
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() - 0.2,
            f'{cnt}篇', ha='center', va='top', fontsize=9, color='white', fontweight='bold')

# 标注最优区间
ax.annotate(
    '最优区间\n+1.05pp',
    xy=(best_idx, ir_values[best_idx]),
    xytext=(best_idx + 0.5, ir_values[best_idx] + 0.8),
    ha='center', fontsize=10, color=COLOR_HIGHLIGHT,
    arrowprops=dict(arrowstyle='->', color=COLOR_HIGHLIGHT, lw=1.5),
    bbox=dict(boxstyle='round,pad=0.3', facecolor='#F0FFF0', edgecolor=COLOR_HIGHLIGHT, alpha=0.8),
)

ax.set_ylabel('平均互动率 (%)', fontsize=12)
ax.set_title('标题字数 × 互动率（倒 U 型）', fontsize=14, fontweight='bold', pad=15)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.tick_params(axis='both', labelsize=11)
ax.set_ylim(0, max(ir_values) + 1.5)

plt.tight_layout()
path = os.path.join(output_dir, '02_title_length.png')
fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
plt.close(fig)
print(f'✓ 图2 已保存: {path}')

# ============================================================
# 图 3：垂类标签 × 体裁 分层分析（控制混淆变量）
# 这张图展示了分析方法的严谨性——不只看到标签有效，而是分层看
# ============================================================
fig, ax = plt.subplots(figsize=(7, 5))

# 定义垂类标签词
VERTICAL_KW = ['asmr', 'ASMR', '教程', '教学', '监督', '免疫向', '助眠', '沉浸式', '一分钟', '1分钟']

def has_vertical(title):
    return any(kw.lower() in title.lower() for kw in VERTICAL_KW)

# 四组数据
groups = {
    '图文\n无标签': [n for n in notes if n['genre'] == '图文' and not has_vertical(n['title'])],
    '图文\n有标签': [n for n in notes if n['genre'] == '图文' and has_vertical(n['title'])],
    '视频\n无标签': [n for n in notes if n['genre'] == '视频' and not has_vertical(n['title'])],
    '视频\n有标签': [n for n in notes if n['genre'] == '视频' and has_vertical(n['title'])],
}

labels = list(groups.keys())
ir_vals = [sum(n['ir'] for n in g) / len(g) if g else 0 for g in groups.values()]
cnts = [len(g) for g in groups.values()]
colors = [COLOR_IMG, COLOR_IMG, COLOR_VID, COLOR_VID]
alphas = [0.5, 1.0, 0.5, 1.0]  # 无标签 = 半透明，有标签 = 全实心

bars = ax.bar(labels, ir_vals, color=colors, alpha=0.9, width=0.55, edgecolor='white', linewidth=0.8)

# 无标签用斜线纹理区分
for bar, alpha_val in zip(bars, alphas):
    if alpha_val < 1:
        bar.set_hatch('///')
        bar.set_alpha(0.6)

for bar, val, cnt in zip(bars, ir_vals, cnts):
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.08,
            f'{val:.2f}%', ha='center', va='bottom', fontsize=12, fontweight='bold')
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() - 0.2,
            f'{cnt}篇', ha='center', va='top', fontsize=9, color='white', fontweight='bold')

# 标注视频内有标签的效果
delta = ir_vals[3] - ir_vals[2]
ax.annotate(
    f'视频内标签效果\n+{delta:.2f}pp',
    xy=(3, ir_vals[3]),
    xytext=(2.7, ir_vals[3] + 1.2),
    ha='center', fontsize=10, color=COLOR_VID,
    arrowprops=dict(arrowstyle='->', color=COLOR_VID, lw=1.5),
    bbox=dict(boxstyle='round,pad=0.3', facecolor='#FFF5EE', edgecolor=COLOR_VID, alpha=0.8),
)

# 图例
from matplotlib.patches import Patch
legend_elements = [
    Patch(facecolor=COLOR_IMG, alpha=0.9, label='图文'),
    Patch(facecolor=COLOR_VID, alpha=0.9, label='视频'),
    Patch(facecolor='white', hatch='///', alpha=0.6, label='无标签(斜线)'),
]
ax.legend(handles=legend_elements, loc='upper left', fontsize=9, framealpha=0.9)

ax.set_ylabel('平均互动率 (%)', fontsize=12)
ax.set_title('垂类标签效果 × 体裁分层（控制混淆变量）', fontsize=14, fontweight='bold', pad=15)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.set_ylim(0, max(ir_vals) + 2)

plt.tight_layout()
path = os.path.join(output_dir, '03_vertical_label_stratified.png')
fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
plt.close(fig)
print(f'✓ 图3 已保存: {path}')

# ============================================================
# 图 4：发布时间段 × 互动率
# 可操作性结论：什么时间发效果好
# ============================================================
fig, ax = plt.subplots(figsize=(8, 4.5))

def time_period(hour):
    if hour is None:
        return None
    if 6 <= hour < 9:
        return '清晨\n6-9点'
    if 9 <= hour < 12:
        return '上午\n9-12点'
    if 12 <= hour < 14:
        return '午间\n12-14点'
    if 14 <= hour < 18:
        return '下午\n14-18点'
    if 18 <= hour < 21:
        return '傍晚\n18-21点'
    if 21 <= hour < 24:
        return '深夜\n21-24点'
    return '凌晨\n0-6点'

period_order = ['清晨\n6-9点', '上午\n9-12点', '午间\n12-14点', '下午\n14-18点',
                '傍晚\n18-21点', '深夜\n21-24点', '凌晨\n0-6点']

period_data = {}
for n in notes:
    tp = time_period(n['hour'])
    if tp not in period_data:
        period_data[tp] = []
    period_data[tp].append(n)

period_labels = [p for p in period_order if p in period_data and period_data[p]]
period_ir = [sum(n['ir'] for n in period_data[p]) / len(period_data[p]) for p in period_labels]
period_cnts = [len(period_data[p]) for p in period_labels]

# 计算整体均值线
overall_avg = sum(n['ir'] for n in notes) / len(notes)

bars = ax.bar(period_labels, period_ir, color=PALETTE_4[:len(period_labels)],
              width=0.55, edgecolor='white', linewidth=0.8)

# 画整体均值线
ax.axhline(y=overall_avg, color='gray', linestyle='--', linewidth=1.2, alpha=0.7)
ax.text(len(period_labels) - 0.5, overall_avg + 0.05, f'整体均值 {overall_avg:.2f}%',
        fontsize=9, color='gray', ha='right', va='bottom')

for bar, val, cnt in zip(bars, period_ir, period_cnts):
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.05,
            f'{val:.2f}%', ha='center', va='bottom', fontsize=12, fontweight='bold')
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() - 0.15,
            f'{cnt}篇', ha='center', va='top', fontsize=8, color='white', fontweight='bold')

ax.set_ylabel('平均互动率 (%)', fontsize=12)
ax.set_title('发布时间段 × 互动率', fontsize=14, fontweight='bold', pad=15)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.tick_params(axis='both', labelsize=10)
ax.set_ylim(0, max(period_ir) + 1.2)

plt.tight_layout()
path = os.path.join(output_dir, '04_publish_time.png')
fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
plt.close(fig)
print(f'✓ 图4 已保存: {path}')

print(f'\n全部 4 张图表已生成到 {output_dir}/ 目录')
print('可以用文件管理器打开 charts/ 文件夹查看')
