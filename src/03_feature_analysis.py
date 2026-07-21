"""
全量44篇笔记标题特征 vs 互动率 相关性分析
分析4个特征：省略号、字数、垂类标签词、微场景描述
"""

import openpyxl, re, sys, json
from collections import Counter

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except:
        pass

# ── 1. 读取数据 ──
wb = openpyxl.load_workbook(r"D:\桌面\笔记列表明细表\笔记列表明细表.xlsx", data_only=True)
ws = wb.active

notes = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    title = str(row[0]).strip() if row[0] else ''
    genre = str(row[2]).strip() if row[2] else ''
    if not title or not genre:
        continue
    exposure = float(row[3]) if row[3] else 0
    likes = float(row[6]) if row[6] else 0
    comments = float(row[7]) if row[7] else 0
    saves = float(row[8]) if row[8] else 0
    ir = ((likes + saves + comments) / exposure * 100) if exposure > 0 else 0
    notes.append({
        'title': title, 'genre': genre, 'ir': ir, 'exposure': exposure,
        'likes': likes, 'saves': saves, 'comments': comments,
        'publish_time': str(row[1]) if row[1] else '',
    })
wb.close()

print(f"共读取 {len(notes)} 篇有效笔记")

# ── 2. 特征提取 ──

# 特征1: 是否含省略号
def has_ellipsis(title):
    return bool(re.search(r'…|\.\.\.', title))

# 特征2: 字数
def title_length(title):
    return len(title)

# 特征3: 是否含垂类标签词
# 基于对这份数据中出现的品类/形式词的观察
VERTICAL_KEYWORDS = [
    'asmr', 'ASMR', 'Asmr',
    '教程', '教学', '监督',
    '免疫向', '助眠',
    '沉浸式',
    '教程', '攻略', '指南', '方法',
    '测评', '开箱',
    '一分钟', '1分钟',
]
def has_vertical_label(title):
    for kw in VERTICAL_KEYWORDS:
        if kw.lower() in title.lower():
            return True
    return False

# 列出所有命中的标签词，方便检查
def get_vertical_label(title):
    found = []
    for kw in VERTICAL_KEYWORDS:
        if kw.lower() in title.lower():
            found.append(kw)
    return found

# 特征4: 是否描述具象微场景
# 判断逻辑（多信号综合打分）：
#   + 含动作动词 (吃、喝、咬、飞、叫、打、玩、监督、变、进、钻…)
#   + 含拟声/拟态 (囔囔、叽、丝丝、咯咯…)
#   + 含对话/称呼 (人类、师傅、你、我…)
#   + 含具体时刻 (今天、现在、刚…)
#   + 含具体身体部位/物体 (袖筒、手、嘴、豆、芒果…)
#   − 含抽象概括词 (什么情况、怎么回事、为什么、如何…)
#   − 教程/方法类标题
#   − 纯描述性陈述无具体动作

ACTION_VERBS = [
    '吃', '喝', '咬', '飞', '叫', '打', '玩', '进', '钻', '爬', '跳',
    '转', '翻', '监督', '教学', '变', '长', '掉', '换', '跑', '走',
    '喂', '食', '睡', '醒', '动', '抓', '啄', '叼', '碰', '摸', '拍',
    '唱', '演', '做', '弄', '搞', '陪',
]

CONCRETE_NOUNS = [
    '袖筒', '手', '嘴', '豆', '芒果', '牡丹', '手机壳', '球',
    '底盘', '笼子', '毛', '翅膀', '蛋', '水', '奶', '食物',
]

DIALOGUE_MARKERS = [
    '人类', '师傅', '你', '我', '他', '她', '吗', '吧', '呢', '呀', '哦',
]

TIME_MARKERS = [
    '今天', '现在', '刚', '突然', '正要', '一会',
]

ONOMATOPOEIA = [
    '囔囔', '叽', '丝丝', '咯咯', '咕咕', '啾啾', '唧唧',
]

ABSTRACT_PATTERNS = [
    r'什么情况', r'怎么回事', r'为什么', r'如何', r'怎么[样么]',
    r'教程', r'方法', r'攻略', r'指南',
]

def is_micro_scene(title):
    """
    多信号综合判断是否为"具象微场景"标题。
    满足 ≥2 个正面信号 且 不匹配负面信号，判定为微场景。
    """
    score = 0
    details = []

    # 正面信号
    if any(v in title for v in ACTION_VERBS):
        score += 1
        found = [v for v in ACTION_VERBS if v in title]
        details.append(f"动作({','.join(found)})")

    if any(v in title for v in CONCRETE_NOUNS):
        score += 1
        found = [v for v in CONCRETE_NOUNS if v in title]
        details.append(f"具象名词({','.join(found)})")

    if any(v in title for v in DIALOGUE_MARKERS):
        score += 1
        found = [v for v in DIALOGUE_MARKERS if v in title]
        details.append(f"对话标记({','.join(found)})")

    if any(v in title for v in TIME_MARKERS):
        score += 1
        details.append("时间锚点")

    if any(v in title for v in ONOMATOPOEIA):
        score += 1
        found = [v for v in ONOMATOPOEIA if v in title]
        details.append(f"拟声词({','.join(found)})")

    # 负面信号：抽象概括
    has_abstract = any(re.search(p, title) for p in ABSTRACT_PATTERNS)
    if has_abstract:
        score -= 2
        details.append("⚠抽象概括")

    # 教程类
    if any(kw in title for kw in ['教程', '攻略', '指南', '方法']):
        score -= 1
        details.append("⚠教程类")

    is_scene = score >= 2
    return is_scene, score, details

# ── 3. 给每篇笔记打特征标签 ──
for n in notes:
    t = n['title']
    n['has_ellipsis'] = has_ellipsis(t)
    n['length'] = title_length(t)
    n['has_vertical'] = has_vertical_label(t)
    n['vertical_labels'] = get_vertical_label(t)
    is_scene, scene_score, scene_details = is_micro_scene(t)
    n['is_micro_scene'] = is_scene
    n['scene_score'] = scene_score
    n['scene_details'] = scene_details

# ── 4. 分组统计 ──
def group_stats(notes_list, group_name, feature_func, label_true, label_false):
    """按二值特征分组，返回两组的统计"""
    group_true = [n for n in notes_list if feature_func(n)]
    group_false = [n for n in notes_list if not feature_func(n)]

    def stats(g):
        if not g:
            return {'count': 0, 'avg_ir': 0, 'min_ir': 0, 'max_ir': 0, 'median_ir': 0}
        irs = sorted([n['ir'] for n in g])
        return {
            'count': len(g),
            'avg_ir': sum(irs) / len(irs),
            'min_ir': min(irs),
            'max_ir': max(irs),
            'median_ir': irs[len(irs)//2],
        }

    return {
        'true': stats(group_true),
        'false': stats(group_false),
        'true_label': label_true,
        'false_label': label_false,
        'true_notes': group_true,
        'false_notes': group_false,
    }

def group_stats_multi(notes_list, buckets, feature_func):
    """按数值特征分组，返回各桶统计"""
    result = []
    for label, condition in buckets:
        g = [n for n in notes_list if condition(n)]
        irs = sorted([n['ir'] for n in g])
        result.append({
            'label': label,
            'count': len(g),
            'avg_ir': sum(irs)/len(irs) if irs else 0,
            'min_ir': min(irs) if irs else 0,
            'max_ir': max(irs) if irs else 0,
            'median_ir': irs[len(irs)//2] if irs else 0,
        })
    return result

# ── 5. 输出报告 ──
output_lines = []

def p(line=""):
    output_lines.append(line)

p("=" * 80)
p("  小红书笔记标题特征 vs 互动率 全量相关性分析")
p("  样本: 全部 44 篇笔记 | 互动率 = (点赞+收藏+评论)/曝光 × 100%")
p("=" * 80)

# ── 特征1: 省略号 ──
p()
p("━" * 80)
p("  特征 1: 标题是否包含省略号 (… 或 ...)")
p("━" * 80)

ellipsis_result = group_stats(notes, '省略号', lambda n: n['has_ellipsis'], '含省略号', '不含省略号')

p(f"\n{'组别':<20} {'篇数':>6} {'平均互动率':>12} {'中位数':>10} {'最低':>10} {'最高':>10}")
p("-" * 70)

for key in ['true', 'false']:
    s = ellipsis_result[key]
    label = ellipsis_result[f'{key}_label']
    p(f"{label:<20} {s['count']:>6} {s['avg_ir']:>10.2f}% {s['median_ir']:>9.2f}% {s['min_ir']:>9.2f}% {s['max_ir']:>9.2f}%")

# 差异
t_s = ellipsis_result['true']
f_s = ellipsis_result['false']
if t_s['count'] > 0 and f_s['count'] > 0:
    diff = t_s['avg_ir'] - f_s['avg_ir']
    p(f"\n→ 含省略号组互动率比不含组 {'高' if diff > 0 else '低'} {abs(diff):.2f} 个百分点")
    p(f"→ 含省略号: {t_s['count']}篇, 不含: {f_s['count']}篇")

# 列出含省略号的所有笔记
p(f"\n含省略号的笔记 ({t_s['count']}篇):")
for n in sorted(ellipsis_result['true_notes'], key=lambda x: -x['ir']):
    p(f"  [{n['genre']}] {n['title'][:50]}  → 互动率 {n['ir']:.2f}%")

# ── 特征2: 字数 ──
p()
p("━" * 80)
p("  特征 2: 标题字数")
p("━" * 80)

lengths_sorted = sorted([n['length'] for n in notes])
p(f"全部笔记字数分布: 最短 {min(lengths_sorted)} 字, 最长 {max(lengths_sorted)} 字, 平均 {sum(lengths_sorted)/len(lengths_sorted):.1f} 字, 中位数 {lengths_sorted[len(lengths_sorted)//2]} 字")

# 按字数分桶: 极短(≤6), 短(7-9), 中(10-13), 长(≥14)
length_buckets = [
    ("极短 ≤6字", lambda n: n['length'] <= 6),
    ("短 7-9字", lambda n: 7 <= n['length'] <= 9),
    ("中 10-13字", lambda n: 10 <= n['length'] <= 13),
    ("长 ≥14字", lambda n: n['length'] >= 14),
]

len_result = group_stats_multi(notes, length_buckets, None)

p(f"\n{'字数区间':<20} {'篇数':>6} {'平均互动率':>12} {'中位数':>10} {'最低':>10} {'最高':>10}")
p("-" * 70)
for b in len_result:
    p(f"{b['label']:<20} {b['count']:>6} {b['avg_ir']:>10.2f}% {b['median_ir']:>9.2f}% {b['min_ir']:>9.2f}% {b['max_ir']:>9.2f}%")

# 列出各区间的笔记标题
for b in len_result:
    notes_in_bucket = [n for n in notes if (
        (b['label'] == '极短 ≤6字' and n['length'] <= 6) or
        (b['label'] == '短 7-9字' and 7 <= n['length'] <= 9) or
        (b['label'] == '中 10-13字' and 10 <= n['length'] <= 13) or
        (b['label'] == '长 ≥14字' and n['length'] >= 14)
    )]
    p(f"\n  [{b['label']}] 笔记列表 ({len(notes_in_bucket)}篇):")
    for n in sorted(notes_in_bucket, key=lambda x: -x['ir']):
        p(f"    {n['title'][:55]}  [{n['genre']}] IR={n['ir']:.2f}%")

# ── 特征3: 垂类标签词 ──
p()
p("━" * 80)
p("  特征 3: 是否包含垂类标签词 (ASMR/教程/教学/免疫向/助眠/沉浸式 等)")
p("━" * 80)

vertical_result = group_stats(notes, '垂类标签', lambda n: n['has_vertical'], '含垂类标签', '不含垂类标签')

p(f"\n{'组别':<20} {'篇数':>6} {'平均互动率':>12} {'中位数':>10} {'最低':>10} {'最高':>10}")
p("-" * 70)

for key in ['true', 'false']:
    s = vertical_result[key]
    label = vertical_result[f'{key}_label']
    p(f"{label:<20} {s['count']:>6} {s['avg_ir']:>10.2f}% {s['median_ir']:>9.2f}% {s['min_ir']:>9.2f}% {s['max_ir']:>9.2f}%")

vt_s = vertical_result['true']
vf_s = vertical_result['false']
if vt_s['count'] > 0 and vf_s['count'] > 0:
    diff = vt_s['avg_ir'] - vf_s['avg_ir']
    p(f"\n→ 含垂类标签组互动率比不含组 {'高' if diff > 0 else '低'} {abs(diff):.2f} 个百分点")

# 列出含垂类标签的笔记
p(f"\n含垂类标签的笔记 ({vt_s['count']}篇):")
for n in sorted(vertical_result['true_notes'], key=lambda x: -x['ir']):
    p(f"  [{n['genre']}] {n['title'][:50]}  → 互动率 {n['ir']:.2f}%  标签: {', '.join(n['vertical_labels'])}")

# ── 特征4: 微场景 ──
p()
p("━" * 80)
p("  特征 4: 是否描述具象微场景 (含动作/对话/拟声/具体时刻, 排除抽象概括)")
p("━" * 80)
p("  判定规则: 动作动词/具象名词/对话标记/时间锚点/拟声词 ≥2个正面信号, 且不匹配抽象概括模式")

scene_result = group_stats(notes, '微场景', lambda n: n['is_micro_scene'], '微场景描述', '笼统描述')

p(f"\n{'组别':<20} {'篇数':>6} {'平均互动率':>12} {'中位数':>10} {'最低':>10} {'最高':>10}")
p("-" * 70)

for key in ['true', 'false']:
    s = scene_result[key]
    label = scene_result[f'{key}_label']
    p(f"{label:<20} {s['count']:>6} {s['avg_ir']:>10.2f}% {s['median_ir']:>9.2f}% {s['min_ir']:>9.2f}% {s['max_ir']:>9.2f}%")

st_s = scene_result['true']
sf_s = scene_result['false']
if st_s['count'] > 0 and sf_s['count'] > 0:
    diff = st_s['avg_ir'] - sf_s['avg_ir']
    p(f"\n→ 微场景组互动率比笼统组 {'高' if diff > 0 else '低'} {abs(diff):.2f} 个百分点")

# 列出微场景笔记
p(f"\n微场景描述笔记 ({st_s['count']}篇):")
for n in sorted(scene_result['true_notes'], key=lambda x: -x['ir']):
    signals = ', '.join(n['scene_details']) if n['scene_details'] else '(无信号)'
    p(f"  [{n['genre']}] {n['title'][:55]}  → IR={n['ir']:.2f}%  [{signals}]")

# 列出笼统描述笔记
p(f"\n笼统描述笔记 ({sf_s['count']}篇):")
for n in sorted(scene_result['false_notes'], key=lambda x: -x['ir']):
    signals = ', '.join(n['scene_details']) if n['scene_details'] else '(无信号)'
    p(f"  [{n['genre']}] {n['title'][:55]}  → IR={n['ir']:.2f}%  [{signals}]")

# ── 6. 组合特征分析 ──
p()
p("━" * 80)
p("  组合特征: 两项最强特征的叠加效应")
p("━" * 80)

# 省略号 × 微场景
combo_groups = [
    ("含省略号 + 微场景", lambda n: n['has_ellipsis'] and n['is_micro_scene']),
    ("含省略号 + 非微场景", lambda n: n['has_ellipsis'] and not n['is_micro_scene']),
    ("无省略号 + 微场景", lambda n: not n['has_ellipsis'] and n['is_micro_scene']),
    ("无省略号 + 非微场景", lambda n: not n['has_ellipsis'] and not n['is_micro_scene']),
]

p(f"\n{'组合':<25} {'篇数':>6} {'平均互动率':>12} {'中位数':>10}")
p("-" * 58)
for label, cond in combo_groups:
    g = [n for n in notes if cond(n)]
    irs = sorted([n['ir'] for n in g])
    avg = sum(irs)/len(irs) if irs else 0
    med = irs[len(irs)//2] if irs else 0
    p(f"{label:<25} {len(g):>6} {avg:>10.2f}% {med:>9.2f}%")

# 垂类标签 × 微场景
p()
p(f"\n{'组合':<25} {'篇数':>6} {'平均互动率':>12} {'中位数':>10}")
p("-" * 58)
combo_groups2 = [
    ("垂类标签 + 微场景", lambda n: n['has_vertical'] and n['is_micro_scene']),
    ("垂类标签 + 非微场景", lambda n: n['has_vertical'] and not n['is_micro_scene']),
    ("无垂类标签 + 微场景", lambda n: not n['has_vertical'] and n['is_micro_scene']),
    ("无垂类标签 + 非微场景", lambda n: not n['has_vertical'] and not n['is_micro_scene']),
]
for label, cond in combo_groups2:
    g = [n for n in notes if cond(n)]
    irs = sorted([n['ir'] for n in g])
    avg = sum(irs)/len(irs) if irs else 0
    med = irs[len(irs)//2] if irs else 0
    p(f"{label:<25} {len(g):>6} {avg:>10.2f}% {med:>9.2f}%")

# ── 7. 全量笔记一览表 ──
p()
p("━" * 80)
p("  附录: 全部 44 篇笔记特征标注一览 (按互动率降序)")
p("━" * 80)

p(f"\n{'排名':<4} {'标题':<40} {'IR%':>6} {'省略':>4} {'字数':>4} {'标签':>4} {'微场景':>6} {'体裁':>4}")
p("-" * 78)

for i, n in enumerate(sorted(notes, key=lambda x: -x['ir']), 1):
    title_short = n['title'][:38] + ".." if len(n['title']) > 38 else n['title']
    ell = "✓" if n['has_ellipsis'] else "✗"
    vert = "✓" if n['has_vertical'] else "✗"
    scene = "✓" if n['is_micro_scene'] else "✗"
    p(f"{i:<4} {title_short:<40} {n['ir']:>5.2f}% {ell:>4} {n['length']:>4} {vert:>4} {scene:>6} {n['genre']:>4}")

# ── 8. 总结 ──
p()
p("=" * 80)
p("  总结")
p("=" * 80)

p(f"""
【样本说明】全部 44 篇笔记: 图文 32 篇, 视频 12 篇
互动率 = (点赞 + 收藏 + 评论) / 曝光 × 100%

【四个特征的单变量效果量 (effect size)】

特征1 - 省略号:
  含省略号 {t_s['count']}篇, 平均互动率 {t_s['avg_ir']:.2f}%
  不含省略号 {f_s['count']}篇, 平均互动率 {f_s['avg_ir']:.2f}%
  差异: {t_s['avg_ir'] - f_s['avg_ir']:+.2f} 个百分点

特征2 - 字数:
  按字数分桶的互动率差异见上文表格
  字数与互动率的 Pearson 相关系数需要额外计算

特征3 - 垂类标签:
  含垂类标签 {vt_s['count']}篇, 平均互动率 {vt_s['avg_ir']:.2f}%
  不含垂类标签 {vf_s['count']}篇, 平均互动率 {vf_s['avg_ir']:.2f}%
  差异: {vt_s['avg_ir'] - vf_s['avg_ir']:+.2f} 个百分点

特征4 - 微场景:
  微场景描述 {st_s['count']}篇, 平均互动率 {st_s['avg_ir']:.2f}%
  笼统描述 {sf_s['count']}篇, 平均互动率 {sf_s['avg_ir']:.2f}%
  差异: {st_s['avg_ir'] - sf_s['avg_ir']:+.2f} 个百分点

【方法论说明】
- "省略号": 标题包含 "…" 或 "..." 即判定为含省略号
- "垂类标签词": 基于词库匹配, 包含 ASMR/教程/教学/免疫向/助眠/沉浸式 等品类形式词
- "微场景": 多信号综合打分 —
    正面信号: 动作动词、具象名词、对话标记(你/我/人类/吗/吧等)、
              时间锚点(今天/现在/突然等)、拟声词
    负面信号: 抽象概括模式(什么情况/如何/怎么/为什么)、教程类标题
    判定阈值: ≥2 个正面信号且未被负面信号抵消
""")

# ── 写入文件 ──
result = '\n'.join(output_lines)
out_path = r"D:\桌面\笔记列表明细表\full_title_feature_analysis.txt"
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(result)

print(f"\n分析完成，结果已写入: {out_path}")
print(result)
