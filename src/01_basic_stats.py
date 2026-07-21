"""
小红书笔记数据分析脚本
读取 笔记列表明细表.xlsx，分析图文/视频两种体裁的互动率，并列出前10篇笔记。
"""

import openpyxl
import sys
import os

# 修复 Windows 终端 GBK 编码问题
if sys.platform == "win32":
    os.system("chcp 65001 > nul")  # 切换到 UTF-8 代码页
    # 或者用 sys.stdout.reconfigure
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass


def main():
    filepath = r"D:\桌面\笔记列表明细表\笔记列表明细表.xlsx"

    # --- 1. 读取数据 ---
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # 真正的表头在第二行（第1行是总标题"达人笔记前1000条笔记"），数据从第3行开始
    headers = [cell.value for cell in list(ws.iter_rows(min_row=2, max_row=2))[0]]
    print("检测到的表头:", headers, "\n")

    # 按列名定位（用位置更稳健，因为表头可能有空格或不可见字符）
    # 列顺序: 0笔记标题, 1首次发布时间, 2体裁, 3曝光, 4观看量, 5封面点击率,
    #          6点赞, 7评论, 8收藏, 9涨粉, 10分享, 11人均观看时长, 12弹幕
    COL_TITLE = 0
    COL_GENRE = 2
    COL_EXPOSURE = 3
    COL_VIEWS = 4
    COL_CLICK_RATE = 5
    COL_LIKES = 6
    COL_COMMENTS = 7
    COL_SAVES = 8
    COL_FOLLOWERS = 9
    COL_SHARES = 10
    COL_WATCH_TIME = 11

    notes = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        title = str(row[COL_TITLE]).strip() if row[COL_TITLE] else ""
        genre = str(row[COL_GENRE]).strip() if row[COL_GENRE] else ""
        if not title or not genre:
            continue

        exposure = float(row[COL_EXPOSURE]) if row[COL_EXPOSURE] else 0
        views = float(row[COL_VIEWS]) if row[COL_VIEWS] else 0
        likes = float(row[COL_LIKES]) if row[COL_LIKES] else 0
        comments = float(row[COL_COMMENTS]) if row[COL_COMMENTS] else 0
        saves = float(row[COL_SAVES]) if row[COL_SAVES] else 0
        shares = float(row[COL_SHARES]) if row[COL_SHARES] else 0
        followers = float(row[COL_FOLLOWERS]) if row[COL_FOLLOWERS] else 0
        watch_time = float(row[COL_WATCH_TIME]) if row[COL_WATCH_TIME] else 0
        click_rate = float(row[COL_CLICK_RATE]) if row[COL_CLICK_RATE] else 0

        # 计算各项互动率（曝光为0则互动率为0，避免除零）
        like_rate = (likes / exposure * 100) if exposure > 0 else 0
        save_rate = (saves / exposure * 100) if exposure > 0 else 0
        comment_rate = (comments / exposure * 100) if exposure > 0 else 0
        # 综合互动率 = (点赞 + 收藏 + 评论) / 曝光
        interaction_rate = ((likes + saves + comments) / exposure * 100) if exposure > 0 else 0

        notes.append({
            "title": title,
            "genre": genre,
            "exposure": exposure,
            "views": views,
            "likes": likes,
            "comments": comments,
            "saves": saves,
            "shares": shares,
            "followers": followers,
            "watch_time": watch_time,
            "click_rate": click_rate,
            "like_rate": like_rate,
            "save_rate": save_rate,
            "comment_rate": comment_rate,
            "interaction_rate": interaction_rate,
            "publish_time": str(row[COL_TITLE + 1]) if row[COL_TITLE + 1] else "",
        })

    wb.close()

    # --- 2. 按体裁分组计算平均互动率 ---
    img_text_notes = [n for n in notes if n["genre"] == "图文"]
    video_notes = [n for n in notes if n["genre"] == "视频"]

    def avg_rate(note_list, key):
        if not note_list:
            return 0
        return sum(n[key] for n in note_list) / len(note_list)

    print("=" * 70)
    print("体裁平均互动率对比")
    print("=" * 70)

    for label, group in [("图文", img_text_notes), ("视频", video_notes)]:
        if not group:
            print(f"\n{label}: 无数据")
            continue
        avg_like = avg_rate(group, "like_rate")
        avg_save = avg_rate(group, "save_rate")
        avg_comment = avg_rate(group, "comment_rate")
        avg_interact = avg_rate(group, "interaction_rate")
        print(f"\n[{label}] ({len(group)} 篇笔记)")
        print(f"   平均点赞率 (点赞/曝光): {avg_like:.2f}%")
        print(f"   平均收藏率 (收藏/曝光): {avg_save:.2f}%")
        print(f"   平均评论率 (评论/曝光): {avg_comment:.2f}%")
        print(f"   平均综合互动率 ((点赞+收藏+评论)/曝光): {avg_interact:.2f}%")

    # --- 3. 按综合互动率排名，列出前10 ---
    print("\n" + "=" * 70)
    print("综合互动率 TOP 10 笔记 (按 (点赞+收藏+评论)/曝光 排名)")
    print("=" * 70)

    ranked = sorted(notes, key=lambda n: n["interaction_rate"], reverse=True)
    top10 = ranked[:10]

    print(f"\n{'排名':<4} {'标题':<35} {'体裁':<6} {'曝光':>10} {'点赞':>8} {'收藏':>8} {'评论':>8} {'互动率':>8}")
    print("-" * 95)

    for i, n in enumerate(top10, 1):
        title_display = n["title"][:32] + "..." if len(n["title"]) > 32 else n["title"]
        print(f"{i:<4} {title_display:<35} {n['genre']:<6} "
              f"{n['exposure']:>10.0f} {n['likes']:>8.0f} {n['saves']:>8.0f} {n['comments']:>8.0f} "
              f"{n['interaction_rate']:>7.2f}%")

    # 也展示按点赞率的 TOP 10
    print("\n" + "=" * 70)
    print("点赞率 TOP 10 笔记 (按 点赞/曝光 排名)")
    print("=" * 70)

    ranked_like = sorted(notes, key=lambda n: n["like_rate"], reverse=True)
    top10_like = ranked_like[:10]

    print(f"\n{'排名':<4} {'标题':<35} {'体裁':<6} {'曝光':>10} {'点赞':>8} {'点赞率':>8}")
    print("-" * 75)

    for i, n in enumerate(top10_like, 1):
        title_display = n["title"][:32] + "..." if len(n["title"]) > 32 else n["title"]
        print(f"{i:<4} {title_display:<35} {n['genre']:<6} "
              f"{n['exposure']:>10.0f} {n['likes']:>8.0f} "
              f"{n['like_rate']:>7.2f}%")

    # 按收藏率的 TOP 10
    print("\n" + "=" * 70)
    print("收藏率 TOP 10 笔记 (按 收藏/曝光 排名)")
    print("=" * 70)

    ranked_save = sorted(notes, key=lambda n: n["save_rate"], reverse=True)
    top10_save = ranked_save[:10]

    print(f"\n{'排名':<4} {'标题':<35} {'体裁':<6} {'曝光':>10} {'收藏':>8} {'收藏率':>8}")
    print("-" * 75)

    for i, n in enumerate(top10_save, 1):
        title_display = n["title"][:32] + "..." if len(n["title"]) > 32 else n["title"]
        print(f"{i:<4} {title_display:<35} {n['genre']:<6} "
              f"{n['exposure']:>10.0f} {n['saves']:>8.0f} "
              f"{n['save_rate']:>7.2f}%")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)
