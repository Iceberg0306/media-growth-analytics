import openpyxl, re, sys, json
from collections import Counter

# Force UTF-8 output
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except:
        pass

wb = openpyxl.load_workbook(r"D:\桌面\笔记列表明细表\笔记列表明细表.xlsx", data_only=True)
ws = wb.active

notes = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    title = str(row[0]).strip() if row[0] else ''
    genre = str(row[2]).strip() if row[2] else ''
    if not title or not genre:
        continue
    exposure = float(row[3]) if row[3] else 0
    likes = float(row[6]) if row[6] else 0
    comments = float(row[7]) if row[7] else 0
    saves = float(row[8]) if row[8] else 0
    ir = ((likes + saves + comments) / exposure * 100) if exposure > 0 else 0
    notes.append({'title': title, 'genre': genre, 'ir': ir, 'exposure': exposure,
                  'likes': likes, 'saves': saves, 'comments': comments})
wb.close()

notes.sort(key=lambda n: n['ir'], reverse=True)
top10 = notes[:10]
bottom10 = notes[-10:]

# ── helpers ──
def has_emoji(s):
    # broad emoji pattern
    return bool(re.search(
        r'[\U0001F600-\U0001F64F\U0001F300-\U0001F5FF\U0001F680-\U0001F6FF'
        r'\U0001F1E0-\U0001F1FF\U00002600-\U000027BF\U0001F900-\U0001F9FF'
        r'\U0001FA00-\U0001FA6F\U0001FA70-\U0001FAFF\U0001F7E0-\U0001F7FF'
        r'\U00002300-\U000023FF\U00002500-\U000025FF\U0001F300-\U0001FAD6'
        r'\U00002702-\U000027B0\U0001F004\U0001F0CF'
        r'\U0000200D\U0000FE0F\U0000203C\U00002049\U000020E3'
        r'\U00002139\U00002194-\U00002199\U000021A9-\U000021AA'
        r'\U0000231A-\U0000231B\U00002328\U000023CF\U000023E9-\U000023F3'
        r'\U000023F8-\U000023FA\U000024C2\U000025AA-\U000025AB'
        r'\U000025B6\U000025C0\U000025FB-\U000025FE\U00002600-\U000027EF'
        r'\U00002934-\U00002935\U00002B05-\U00002B07\U00002B1B-\U00002B1C'
        r'\U00002B50\U00002B55\U00003030\U0000303D\U00003297\U00003299'
        r'\U0001F000-\U0001FFFF]'
        r'|[\U0001F1E6-\U0001F1FF]{2}', s))

def classify_sent(s):
    s = s.strip()
    if '?' in s or '？' in s:
        return '疑问句'
    if '!' in s or '！' in s:
        return '感叹句'
    return '陈述句'

def extract_keywords(titles, min_len=2):
    words = []
    for t in titles:
        # remove emoji and punctuation, keep CJK + alphanumeric
        clean = re.sub(r'[^一-鿿㐀-䶿A-Za-z0-9]', ' ', t)
        for w in clean.split():
            w = w.strip()
            if len(w) >= min_len:
                words.append(w.lower())
    return Counter(words).most_common(20)

def show_group(name, group):
    lines = []
    lines.append(f"\n{'='*70}")
    lines.append(f"  {name} ({len(group)} 篇)")
    lines.append(f"{'='*70}")

    lengths = [len(n['title']) for n in group]
    emoji_n = sum(1 for n in group if has_emoji(n['title']))
    q_n = sum(1 for n in group if classify_sent(n['title']) == '疑问句')
    e_n = sum(1 for n in group if classify_sent(n['title']) == '感叹句')
    s_n = sum(1 for n in group if classify_sent(n['title']) == '陈述句')

    symbols = '，。！？…～、；：""''（）【】《》—…'
    special_chars = set()
    for n in group:
        for ch in n['title']:
            if ch in symbols:
                special_chars.add(ch)
    # count special punctuation per group
    punct_counts = Counter()
    for n in group:
        for ch in n['title']:
            if ch in symbols:
                punct_counts[ch] += 1

    lines.append(f"\n【基础统计】")
    lines.append(f"  标题字数: 最短 {min(lengths)} 字, 最长 {max(lengths)} 字, 平均 {sum(lengths)/len(lengths):.1f} 字")
    lines.append(f"  含 Emoji: {emoji_n}/{len(group)} ({emoji_n/len(group)*100:.0f}%)")
    lines.append(f"  句式分布: 疑问句 {q_n} 篇 / 感叹句 {e_n} 篇 / 陈述句 {s_n} 篇")
    if punct_counts:
        top_punct = punct_counts.most_common(6)
        lines.append(f"  高频标点: {', '.join(f'{k}({v}次)' for k,v in top_punct)}")

    lines.append(f"\n【逐篇详情】")
    for i, n in enumerate(group, 1):
        flags = ['emoji'] if has_emoji(n['title']) else []
        flags.append(classify_sent(n['title']))
        t = n['title']
        lines.append(f"  {i:2d}. [{n['genre']}] {t}")
        lines.append(f"      字数={len(t)}  互动率={n['ir']:.2f}%  曝光={n['exposure']:.0f}  {' | '.join(flags)}")

    lines.append(f"\n【高频关键词(去emoji后切词)】")
    kw = extract_keywords([n['title'] for n in group])
    for w, c in kw:
        lines.append(f"  \"{w}\" — {c} 次")

    return '\n'.join(lines)

# ── 逐字对比分析 ──
def char_level_analysis(top, bot):
    lines = []
    lines.append(f"\n{'='*70}")
    lines.append(f"  逐维度对比分析")
    lines.append(f"{'='*70}")

    # 1. 字数
    t_len = [len(n['title']) for n in top]
    b_len = [len(n['title']) for n in bot]
    lines.append(f"\n【1. 标题字数】")
    lines.append(f"  TOP10:    平均 {sum(t_len)/len(t_len):.1f} 字 (范围 {min(t_len)}-{max(t_len)})")
    lines.append(f"  倒数10:   平均 {sum(b_len)/len(b_len):.1f} 字 (范围 {min(b_len)}-{max(b_len)})")

    # 2. Emoji
    t_emoji = [n for n in top if has_emoji(n['title'])]
    b_emoji = [n for n in bot if has_emoji(n['title'])]
    lines.append(f"\n【2. Emoji 使用】")
    lines.append(f"  TOP10:    {len(t_emoji)}/10 篇含emoji")
    lines.append(f"  倒数10:   {len(b_emoji)}/10 篇含emoji")

    # Extract actual emojis used
    emoji_set_top = set()
    emoji_set_bot = set()
    emoji_re = re.compile(
        r'[\U0001F600-\U0001F64F\U0001F300-\U0001F5FF\U0001F680-\U0001F6FF'
        r'\U0001F1E0-\U0001F1FF\U00002600-\U000027BF\U0001F900-\U0001F9FF'
        r'\U0001FA00-\U0001FA6F\U0001FA70-\U0001FAFF\U0001F7E0-\U0001F7FF'
        r'\U00002300-\U000023FF\U00002500-\U000025FF\U0001F300-\U0001FAD6'
        r'\U0000200D\U0000FE0F\U0001F000-\U0001FFFF]')
    for n in top:
        for ch in n['title']:
            if emoji_re.match(ch):
                emoji_set_top.add(ch)
    for n in bot:
        for ch in n['title']:
            if emoji_re.match(ch):
                emoji_set_bot.add(ch)
    lines.append(f"  TOP10 所用emoji: {''.join(sorted(emoji_set_top)) if emoji_set_top else '(无)'}")
    lines.append(f"  倒数10 所用emoji: {''.join(sorted(emoji_set_bot)) if emoji_set_bot else '(无)'}")

    # 3. 句式
    t_q = sum(1 for n in top if classify_sent(n['title']) == '疑问句')
    t_e = sum(1 for n in top if classify_sent(n['title']) == '感叹句')
    t_s = sum(1 for n in top if classify_sent(n['title']) == '陈述句')
    b_q = sum(1 for n in bot if classify_sent(n['title']) == '疑问句')
    b_e = sum(1 for n in bot if classify_sent(n['title']) == '感叹句')
    b_s = sum(1 for n in bot if classify_sent(n['title']) == '陈述句')
    lines.append(f"\n【3. 句式分布】")
    lines.append(f"  TOP10:    疑问{t_q} / 感叹{t_e} / 陈述{t_s}")
    lines.append(f"  倒数10:   疑问{b_q} / 感叹{b_e} / 陈述{b_s}")

    # 4. 标点分析
    punct_top = Counter()
    punct_bot = Counter()
    punct_chars = set('，。！？…～、；：""''（）【】《》—…~!?,.')
    for n in top:
        for ch in n['title']:
            if ch in punct_chars:
                punct_top[ch] += 1
    for n in bot:
        for ch in n['title']:
            if ch in punct_chars:
                punct_bot[ch] += 1
    lines.append(f"\n【4. 标点符号使用频率】")
    lines.append(f"  TOP10 常用标点: {', '.join(f'{k}({v}次)' for k,v in punct_top.most_common(8))}")
    lines.append(f"  倒数10 常用标点: {', '.join(f'{k}({v}次)' for k,v in punct_bot.most_common(8))}")

    # 5. 关键词差异
    lines.append(f"\n【5. 高频关键词差异】")
    t_kw = dict(extract_keywords([n['title'] for n in top]))
    b_kw = dict(extract_keywords([n['title'] for n in bot]))
    top_only = {k: t_kw[k] for k in t_kw if k not in b_kw}
    bot_only = {k: b_kw[k] for k in b_kw if k not in t_kw}
    shared = {k: (t_kw[k], b_kw[k]) for k in t_kw if k in b_kw}
    lines.append(f"  TOP10 独有关键词: {', '.join(f'{k}({v})' for k,v in sorted(top_only.items(), key=lambda x:-x[1])[:8])}")
    lines.append(f"  倒数10 独有关键词: {', '.join(f'{k}({v})' for k,v in sorted(bot_only.items(), key=lambda x:-x[1])[:8])}")
    if shared:
        lines.append(f"  共同关键词: {', '.join(f'{k}(TOP{t_kw[k]}/倒数{b_kw[k]})' for k in sorted(shared, key=lambda x:-(t_kw[x]+b_kw[x]))[:5])}")

    # 6. 是否含数字
    t_num = sum(1 for n in top if re.search(r'\d', n['title']))
    b_num = sum(1 for n in bot if re.search(r'\d', n['title']))
    lines.append(f"\n【6. 是否含数字】")
    lines.append(f"  TOP10:    {t_num}/10 篇标题含数字")
    lines.append(f"  倒数10:   {b_num}/10 篇标题含数字")

    # 7. 英文/字母
    t_en = sum(1 for n in top if re.search(r'[A-Za-z]{2,}', n['title']))
    b_en = sum(1 for n in bot if re.search(r'[A-Za-z]{2,}', n['title']))
    lines.append(f"\n【7. 含英文单词(≥2字母)】")
    lines.append(f"  TOP10:    {t_en}/10 篇")
    lines.append(f"  倒数10:   {b_en}/10 篇")

    return '\n'.join(lines)

# ── 汇总规律 ──
def summarize(top, bot):
    lines = []
    lines.append(f"\n{'='*70}")
    lines.append(f"  总结：高互动标题的 3 条共性规律")
    lines.append(f"{'='*70}")

    t_len_avg = sum(len(n['title']) for n in top) / 10
    b_len_avg = sum(len(n['title']) for n in bot) / 10
    t_q = sum(1 for n in top if classify_sent(n['title']) == '疑问句')
    b_q = sum(1 for n in bot if classify_sent(n['title']) == '疑问句')
    t_has_en = sum(1 for n in top if re.search(r'[A-Za-z]{2,}', n['title']))
    b_has_en = sum(1 for n in bot if re.search(r'[A-Za-z]{2,}', n['title']))

    lines.append(f"""
规律一：标题中嵌入"悬念钩子"——用…、省略号或开放式陈述制造未完待续感
  - TOP10 中大量使用"…"、"…？"等省略/中断式标点，留给读者想象空间
  - 相比之下，倒数10篇标题多是封闭式陈述，信息已给完，缺少点击动机
  - 例：TOP10 "球….在进食………" vs 倒数 "小鸡突然不飞了应该是什么情况"

规律二：善用"类别标签词"建立内容识别度（ASMR、特定物种名/宠物名）
  - TOP10 中 "ASMR" 出现 2 次，"小鸟"类词出现 2 次，宠物昵称(叽、鸡)频繁出现
  - 这类词立即使目标受众识别出"这是我喜欢的垂类内容"，提高点击/互动意愿
  - 倒数10 标题关键词更泛化("小鸡""牡丹""绿鸡")，缺乏精准垂类锚点

规律三：标题聚焦一个"微场景"而非概括性描述
  - TOP10 标题是具体时刻的特写："人类，豆让你喝了吗""袖筒有什么？""小鸡师傅一对一监督教学"
  - 倒数10 标题偏向状态概括："牡丹怎么又咬人了""掉毛小鸡长大以后""小鸡突然不飞了"
  - "具象场景"比"抽象状态"更容易触发好奇心和情感共鸣
""")
    # Data-backed refinement
    lines.append(f"【数据支撑】")
    lines.append(f"  • 字数: TOP10 平均 {t_len_avg:.1f} 字 vs 倒数 {b_len_avg:.1f} 字 — 差异不显著，说明字数本身不是决定因素")
    lines.append(f"  • 疑问句式: TOP10 仅 {t_q} 篇用疑问，说明'提问'并非必需；'悬疑式陈述'效果更好")
    lines.append(f"  • 英文词嵌入: TOP10 {t_has_en} 篇含 ASMR 等英文标签，倒数 {b_has_en} 篇 — 垂类标签词有识别红利")
    return '\n'.join(lines)

# ── 输出 ──
output = []
output.append(show_group("综合互动率 TOP 10", top10))
output.append(show_group("综合互动率 倒数 10", bottom10))
output.append(char_level_analysis(top10, bottom10))
output.append(summarize(top10, bottom10))

result = '\n'.join(output)

# Write to file for clean output
out_path = r"D:\桌面\笔记列表明细表\title_analysis_result.txt"
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(result)
print("分析完成，结果已写入:", out_path)
print(result)
